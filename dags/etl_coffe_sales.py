# /opt/airflow/dags/coffee_sales_etl.py
import os
import hashlib
import re
from datetime import datetime, timezone, timedelta
from typing import Optional
from openpyxl import load_workbook
from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.providers.postgres.operators.postgres import PostgresOperator
from airflow.exceptions import AirflowException
from airflow.models import Variable

# ---------- Конфигурация ----------
FILE_PATH = Variable.get("COFFEE_SALES_FILE_PATH", default_var="/opt/airflow/data/coffee_sales.xlsx")
SRC_NAME = os.path.basename(FILE_PATH)

default_args = {
    "owner": "data_engineer",
    "retries": 2,
    "retry_delay": timedelta(minutes=1),
}

# --- Вспомогательные функции (без изменений) ---
def normalize_null(val):
    if val is None:
        return None
    s = str(val).strip()
    return None if s.lower() in ("null", "none", "", "nan") else s

def parse_datetime(dt_str):
    if not dt_str:
        return None
    s = normalize_null(dt_str)
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S.%f").replace(tzinfo=timezone.utc)
    except ValueError:
        return None

def parse_date(d_str):
    if not d_str:
        return None
    s = normalize_null(d_str)
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S").date()
    except ValueError:
        return None

def parse_money(money_str):
    if not money_str:
        return None
    s = normalize_null(money_str)
    if not s:
        return None
    cleaned = re.sub(r"[^0-9.,]", "", s)
    if not cleaned:
        return None
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None

def parse_int(val):
    s = normalize_null(val)
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None

def compute_row_hash(row_dict: dict):
    items = []
    for k, v in sorted(row_dict.items()):
        if v is None:
            items.append(f"{k}=NULL")
        else:
            items.append(f"{k}={v}")
    return hashlib.md5("".join(items).encode("utf-8")).hexdigest()

# --- загрузка в stg ---
def extract_excel_to_stg(**context):
    pg_hook = PostgresHook(postgres_conn_id="postgres_default")
    conn = pg_hook.get_conn()
    cursor = conn.cursor()

    # 1. Определяем date_from и date_to
    params = context["params"] or {}
    logical_date = context["logical_date"].replace(tzinfo=timezone.utc)  # DAG run time (UTC)

    # Если параметры не заданы — вычисляем
    if not params.get("date_from") or not params.get("date_to"):
        # Получаем MAX(sale_datetime) из stg для этого источника
        cursor.execute(
            "SELECT MAX(sale_datetime) FROM stg.coffee_sales WHERE src_name = %s",
            (SRC_NAME,)
        )
        max_dt_row = cursor.fetchone()
        max_dt = max_dt_row[0] if max_dt_row and max_dt_row[0] else datetime(2024, 1, 1, tzinfo=timezone.utc)

        date_from = max_dt
        date_to = logical_date  # не включая — фильтр: >= date_from AND < date_to
        mode = "auto-incremental"
    else:
        # Явно заданы — парсим
        try:
            date_from = datetime.fromisoformat(params["date_from"]).replace(tzinfo=timezone.utc)
            date_to = datetime.fromisoformat(params["date_to"]).replace(tzinfo=timezone.utc)
        except Exception as e:
            raise AirflowException(f"Ошибка парсинга дат: {e}")
        mode = "manual"

    print(f"Режим: {mode} | Диапазон: [{date_from.isoformat()}, {date_to.isoformat()})")
    print(f"SRC_NAME: {SRC_NAME}, FILE_PATH: {FILE_PATH}")

    # 2. Очистка stg: удаляем всё за [date_from, date_to) для этого источника
    delete_sql = """
        DELETE FROM stg.coffee_sales
        WHERE src_name = %s
          AND sale_datetime >= %s
          AND sale_datetime < %s;  -- строгое неравенство, т.к. date_to = logical_date (не включительно)
    """
    cursor.execute(delete_sql, (SRC_NAME, date_from, date_to))
    deleted = cursor.rowcount
    print(f"Удалено {deleted} строк из stg")

    # 3. Чтение Excel
    wb = load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    expected_headers = [
        "date", "datetime", "hour_of_day", "cash_type", "card",
        "money", "coffee_name", "Time_of_Day", "Weekday", "Month_name",
        "Weekdaysort", "Monthsort"
    ]
    if headers != expected_headers:
        raise AirflowException(f"Несовпадение заголовков: {headers}")

    batch = []
    row_count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, row))

        sale_datetime = parse_datetime(raw["datetime"])
        print(f"sale_datetime: {sale_datetime}")
        print(f'sale_date: {raw["date"]}')
        if not sale_datetime:
            skipped += 1
            continue

        # Фильтр: [date_from, date_to)
        if not (date_from <= sale_datetime < date_to):
            skipped += 1
            continue

        record = {
            "sale_date": parse_date(raw["date"]),
            "sale_datetime": sale_datetime,
            "hour_of_day": parse_int(raw["hour_of_day"]),
            "payment_type": normalize_null(raw["cash_type"]),
            "card_anon_hash": normalize_null(raw["card"]),
            "revenue_rub": parse_money(raw["money"]),
            "coffee_name": normalize_null(raw["coffee_name"]),
            "time_of_day": normalize_null(raw["Time_of_Day"]),
            "weekday": normalize_null(raw["Weekday"]),
            "month_name": normalize_null(raw["Month_name"]),
            "weekdaysort": parse_int(raw["Weekdaysort"]),
            "monthsort": parse_int(raw["Monthsort"]),
        }

        hashable = {k: v for k, v in record.items() if v is not None}
        row_hash = compute_row_hash(hashable)

        record["load_date"] = datetime.now(timezone.utc)
        record["src_name"] = SRC_NAME
        record["row_hash"] = row_hash

        batch.append(record)
        row_count += 1


    if batch:
        _insert_batch_to_stg(cursor, batch)


    conn.commit()
    cursor.close()
    conn.close()

    print(f"Загружено {row_count} строк в stg, пропущено {skipped}")
    context["ti"].xcom_push(key="rows_loaded", value=row_count)
    context["ti"].xcom_push(key="date_from", value=date_from.isoformat())
    context["ti"].xcom_push(key="date_to", value=date_to.isoformat())


def _insert_batch_to_stg(cursor, batch):
    from psycopg2.extras import execute_values
    keys = [
        "sale_date", "sale_datetime", "hour_of_day", "payment_type",
        "card_anon_hash", "revenue_rub", "coffee_name", "time_of_day",
        "weekday", "month_name", "weekdaysort", "monthsort",
        "load_date", "src_name", "row_hash"
    ]
    query = f"""
        INSERT INTO stg.coffee_sales ({', '.join(keys)})
        VALUES %s
        ON CONFLICT (src_name, sale_datetime, row_hash) DO NOTHING;
    """
    values = [
        (
            r["sale_date"], r["sale_datetime"], r["hour_of_day"], r["payment_type"],
            r["card_anon_hash"], r["revenue_rub"], r["coffee_name"], r["time_of_day"],
            r["weekday"], r["month_name"], r["weekdaysort"], r["monthsort"],
            r["load_date"], r["src_name"], r["row_hash"]
        )
        for r in batch
    ]
    execute_values(cursor, query, values)
    print(f"query: {query}")



# ---------- DAG ----------
with DAG(
    dag_id="coffee_sales_etl",
    default_args=default_args,
    description="ETL v3: авто-инкремент (если нет параметров) + ручной диапазон",
    schedule_interval="@daily",  # теперь можно и по расписанию!
    start_date=datetime(2024, 1, 1, tzinfo=timezone.utc),
    catchup=False,
    tags=["etl", "coffee", "incremental-auto"],
) as dag:

    extract_task = PythonOperator(
        task_id="extract_excel_to_stg",
        python_callable=extract_excel_to_stg,
        provide_context=True,
        # Пример ручного запуска:
        # {"date_from": "2024-03-10T00:00:00+00:00", "date_to": "2024-03-13T00:00:00+00:00"}
    )

    load_ods_task = PostgresOperator(
        task_id="load_to_ods",
        postgres_conn_id="postgres_default",
        sql="""
        INSERT INTO ods.coffee_sales_clean (
            sale_date, sale_datetime, hour_of_day, payment_type,
            card_anon_hash, revenue_rub, coffee_name,
            load_date, src_row_hash, src_name
        )
        SELECT
            sale_date,
            sale_datetime,
            hour_of_day,
            LOWER(payment_type) AS payment_type,
            card_anon_hash,
            revenue_rub,
            coffee_name,
            load_date,
            row_hash AS src_row_hash,
            src_name
        FROM stg.coffee_sales stg
        WHERE stg.load_date > COALESCE((SELECT MAX(load_date) FROM ods.coffee_sales_clean), '1970-01-01')
          AND stg.sale_datetime IS NOT NULL
          AND stg.revenue_rub IS NOT NULL
          AND stg.coffee_name IS NOT NULL
          AND LOWER(stg.payment_type) IN ('cash', 'card')
          AND stg.hour_of_day BETWEEN 0 AND 23
        ON CONFLICT (src_row_hash)
        DO UPDATE SET
            sale_date = EXCLUDED.sale_date,
            sale_datetime = EXCLUDED.sale_datetime,
            hour_of_day = EXCLUDED.hour_of_day,
            payment_type = EXCLUDED.payment_type,
            card_anon_hash = EXCLUDED.card_anon_hash,
            revenue_rub = EXCLUDED.revenue_rub,
            coffee_name = EXCLUDED.coffee_name,
            load_date = EXCLUDED.load_date,
            src_name = EXCLUDED.src_name;
        """,
    )

    build_dm_task = PostgresOperator(
        task_id="build_daily_dm",
        postgres_conn_id="postgres_default",
        sql="""
        INSERT INTO dm.coffee_sales_daily AS dm (
            sale_date, total_revenue_rub, avg_check_rub, total_units_sold,
            top_hour, top_hour_sales, loaded_at
        )
            with 
            source_data as (
            select *
            FROM ods.coffee_sales_clean c1
             WHERE load_date > COALESCE((SELECT MAX(loaded_at) FROM dm.coffee_sales_daily csd ), '1970-01-01')
            ),
            
            agg_by_day as (
             select sale_date,
                 sum(src.revenue_rub) as total_revenue_rub,
                 avg(src.revenue_rub) as avg_check_rub,
                 count(*) as total_units_sold  
             from source_data src
             group by sale_date 
            )
            ,
            
            agg_per_hour as ( 
             select sale_date, 
                 hour_of_day,
                 count(*) as amount_per_hour,
                 sum(revenue_rub)  as revenue_per_hour
              from source_data 
              group by sale_date, hour_of_day
              
              ) ,
              
              max_amount_per_hour as (
              select sale_date,
                  hour_of_day,
                  amount_per_hour,
                  row_number() over (partition by sale_date order by amount_per_hour desc ) as rn
              from 
              agg_per_hour
              )
              
              select abd.sale_date,
                  total_revenue_rub, 
                  avg_check_rub, 
                  total_units_sold,
                  hour_of_day as top_hour,
                  amount_per_hour as top_hour_sales,
                  now() as loaded_at
              from agg_by_day as abd
              left join (select * from max_amount_per_hour where rn = 1) as t
              on abd.sale_date = t.sale_date
        ON CONFLICT (sale_date) DO UPDATE SET
            total_revenue_rub = EXCLUDED.total_revenue_rub,
            avg_check_rub = EXCLUDED.avg_check_rub,
            total_units_sold = EXCLUDED.total_units_sold,
            top_hour = EXCLUDED.top_hour,
            top_hour_sales = EXCLUDED.top_hour_sales,
            loaded_at = EXCLUDED.loaded_at;
        """,
    )


    extract_task >> load_ods_task >> build_dm_task